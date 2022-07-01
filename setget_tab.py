#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import openpyxl
def get_host_info():
    hosts = []
    mytables = openpyxl.load_workbook(filename = "info_host.xlsx")
    mysheet = mytables["Sheet1"]
    # mysheet.max_row
    # mysheet.max_column
    for i in range(mysheet.max_row):
        host = []
        for j in range(mysheet.max_column):
            host.append(mysheet[chr(j+65) + str(i+1)].value)
            j = j+1
        hosts.append(host)
    mytables.close()
    return hosts

def set_vm_table(file,sheet,data=[]):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = sheet
    ws1.append(tuple(data))
    wb.save(file)
    return 0
